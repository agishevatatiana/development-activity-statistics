import tkinter as tk
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from tkinter import messagebox 
from tkinter import filedialog
import re

from apiHelper import APIHelper

wb = Workbook()
get_worksheet = wb.active
color_scale_visibility_rule = ColorScaleRule(
    start_type='percentile', start_value=0, start_color='F8A998',
    mid_type='percentile', mid_value=50, mid_color='F8C798',
    end_type='percentile', end_value=100, end_color='67C25A')
get_worksheet.title = 'GitHub activity report'
 
root = tk.Tk()
root.title("Compare github repositories statistics")

frame_inputs=tk.Frame(root,width=400)
frame_inputs.grid(row=0,column=0, sticky=tk.W+tk.E, padx=10)
frame_buttons=tk.Frame(root,width=400)
frame_buttons.grid(row=1,column=0,sticky=tk.E, pady=10, padx=10)
  
# declaring string variable
# for storing name and password
repo1_var=tk.StringVar()
repo2_var=tk.StringVar()
repositories = [repo1_var, repo2_var]

def validate(repo_data):
    has_validation_issue = False
    for value in repo_data.values():
        if 'Error' in str(value):
            has_validation_issue = True
            messagebox.showerror('Error', value)
            break
    return has_validation_issue

def validate_input(repo: str, label_name: str):
    repo_pattern=r"^([a-zA-Z0-9@_\-.]+\/[a-zA-Z0-9_\-.]{1,100})$"
    is_invalid = repo != '' and re.match(repo_pattern, repo) is None
    label_widget = root.nametowidget(label_name)
    if is_invalid:
        label_widget.config(text="Please provide the valid GitHub repo name!")
        return True
    label_widget.config(text="")
    return False


def get_table_data():
    table_data = {}
    col_keys_indexes = {}
    has_validation_issue = False
    #column titles && get in data
    for index, repo_var in enumerate(repositories):
        repo=repo_var.get()
        column_title = repo.split('/', 1)[1]
        get_worksheet.cell(row=2, column=index + 3, value=column_title)

        #get in data
        col_keys_indexes[column_title] = index + 3
        table_data[column_title] = APIHelper().get_github_api_repo_activity_data(repo)

        #validation
        if validate(table_data[column_title]) == True:
            has_validation_issue = True
            break

    if has_validation_issue == True:
        return 'Error'
    else:
        return {
            'col_keys_indexes': col_keys_indexes,
            'table_data': table_data
        }
    
def save_report():
    file_name = filedialog.asksaveasfilename(defaultextension='.xlsx', initialdir='', initialfile='report.xlsx')
    if file_name != '':
        wb.save(file_name)
    root.quit()
    exit()

def create_report():
    row_titles = {
        'unique_authors': 'Unique Authors', 
        'open_pr': 'Open PRs', 
        'merged_pr': 'Merged PRs', 
        'open_issues': 'Open Issues' , 
        'closed_issues': 'Closed Issues',  
        'files_changed': 'Files Changed', 
        'commits_to_main': 'Commits to main', 
        'commits_to_all': 'Commits to all'
    }
    row_keys_indexes = {}
    for index, row_key in enumerate(row_titles.keys()):
        row_keys_indexes[row_key] = index + 3
        get_worksheet.cell(row=index + 3, column=2, value=row_titles[row_key])


    table = get_table_data()

    if table != 'Error':
        table_data = table['table_data']
        col_keys_indexes = table['col_keys_indexes']

        #fill in data
        for col_title, col_index in col_keys_indexes.items():
            for row_title, row_index in row_keys_indexes.items():
                cell_value = table_data[col_title][row_title]
                get_worksheet.cell(row=row_index, column=col_index, value=cell_value)
                    
                #color scale filter for rows red is min, orange in the mid value, green for max value
                get_worksheet.conditional_formatting.add(f'A{row_index}:I{row_index}', color_scale_visibility_rule)
        save_report()
    
def add():
    amount = len(repositories)
    if  amount >= 6:
        messagebox.showwarning('Limits warning', 'The maximum amount of repositories to compare is 6.')
        return
    repo_var=tk.StringVar()
    repositories.append(repo_var)
    build_view()

create_report_btn = tk.Button(frame_buttons,text = 'Create Report', command = create_report, font=('calibre',16,'bold'))
add_repo_btn=tk.Button(frame_buttons,text = 'Add', command = add, font=('calibre',16,'bold'))

#build repositories inputs view
def build_view(): 
  row_counter = 0
  for index, repo_var in enumerate(repositories):
      repo_label = tk.Label(frame_inputs, text = f'https://github.com/', font=('calibre',16, 'normal'))
      validation_label = tk.Label(
          frame_inputs,
          width=56,
          text = '',
          font=('calibre',16, 'normal'),
          foreground='red',
          anchor="w"
        )
      validation_register= root.register(validate_input)
      repo_entry = tk.Entry(
          frame_inputs,
          textvariable = repo_var,
          font=('calibre',16,'normal'),
          width=56,
          validate='focusout', 
          validatecommand=(validation_register, '%P', validation_label)
        )
      repo_label.grid(row=row_counter,column=0,pady=3)
      repo_entry.grid(row=row_counter,column=1,pady=3)
      validation_label.grid(row=row_counter + 1, column=1, pady=3)
      row_counter = row_counter + 2

  #build buttons view
  add_repo_btn.grid(row=len(repositories),column=1)
  create_report_btn.grid(row=len(repositories),column=0)

build_view()

root.mainloop()


